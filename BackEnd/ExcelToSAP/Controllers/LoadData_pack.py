import code
from datetime import datetime
from openpyxl import load_workbook
import configparser
from common.encrypt.codeDecode import cargar_clave, desencriptar_items
import logging
from common.coms.sap_coms import SAPContextManager


class CustomError(Exception):
    """
    Excepción personalizada
    """
    pass


def productTreesProcessing(data):
    """
    Preparamos la lista de diccionarios del excel a una lista de diccionarios procesada para subirla a sap

    :param data: Datos del excel
    :return: lista de diccionarios lista para procesar
    """
    code_act = None
    arr_resp = []
    arr_aux = []
    dict_aux = {}
    # Este es probablemente la peor pieza de codigo que he escrito, pero
    # no se como hacerlo mejor
    for i in data:
        if code_act is None:
            code_act = i['ItemCode_Lote']
            dict_aux = {
                'ItemCode_Lote': code_act,
                'Almacen': i['Almacen'],
                'Canal': i['Canal']
            }
            arr_aux.append({
                'ItemCode': i['Item_Code_Producto'],
                'Almacen': i['Almacen'],
                'Cantidad': i['Cantidad']
            })
        elif code_act == i['ItemCode_Lote']:
            arr_aux.append({
                'ItemCode': i['Item_Code_Producto'],
                'Almacen': i['Almacen'],
                'Cantidad': i['Cantidad']
            })
        else:
            dict_aux['Items'] = arr_aux
            arr_resp.append(dict_aux)
            arr_aux = []
            dict_aux = {}
            code_act = i['ItemCode_Lote']
            dict_aux = {
                'ItemCode_Lote': code_act,
                'Almacen': i['Almacen'],
                'Canal': i['Canal']
            }
            arr_aux.append({
                'ItemCode': i['Item_Code_Producto'],
                'Almacen': i['Almacen'],
                'Cantidad': i['Cantidad']
            })
    dict_aux['Items'] = arr_aux
    arr_resp.append(dict_aux)
    return arr_resp


def excel_to_hanadb_pack(database):
    """
    Cuerpo del programa ExcelToSAP para creación de packs. Abrimos el archivo de excel, leemos los datos y
    actualizamos SAP.

    :param database: Nombre de la base de datos que se quiere actualizar.
    """
    parser = configparser.ConfigParser()
    parser.read('./config_code.ini')
    db = {}
    if parser.has_section('SAP'):
        params = parser.items('SAP')
        for param in params:
            db[param[0]] = (desencriptar_items(param[1],
                                               cargar_clave()).decode(
                "utf-8"))
    workbook = load_workbook('tempExcelToSAPSistemas.xlsx')
    sheets = workbook.sheetnames
    logging.info("Abrimos excel")

    for i in sheets:
        if i == "Items":
            try:
                data_excel = excel_to_dict('tempExcelToSAPSistemas.xlsx', i)
                logging.info(
                    f"Se cargan {len(data_excel)} registros del excel")
                with SAPContextManager(ip=db['ip'], CompanyDB=database,
                                       UserName=db['username'], password=db['password'],
                                       ) as conn:
                    comprobar_excel(data_excel, i)
                    for row in data_excel:
                        if row['ItemCode'] is None:
                            raise (CustomError("El campo ItemCode no puede estar"
                                               " vacío"))
                        logging.info(f"Actualizando {row['ItemCode']}")
                        conn.crearItem(ItemCode=row['ItemCode'],
                                       Nom_Red=row['Nombre_Reducido'],
                                       Descrip=row['Descripcion'],
                                       ean=row['EAN'],
                                       GrupoImp=row['Grupo Impositivo']
                                       )
            except Exception as e:
                return -1, e

        elif i == "ProductTrees":
            try:
                data_excel = excel_to_dict('tempExcelToSAPSistemas.xlsx', i)
                logging.info(
                    f"Se cargan {len(data_excel)} registros del excel")
                with SAPContextManager(ip=db['ip'], CompanyDB=database,
                                       UserName=db['username'], password=db['password'],
                                       ) as conn:
                    comprobar_excel(data_excel, i)
                    list_items = productTreesProcessing(data_excel)
                    for i in list_items:
                        conn.crearProductTree(i)
            except Exception as e:
                return -1, e
        else:
            return -1, f"Revisa el nombre de la hoja {i}"
    return 1, 'Se ha ejecutado con exito'


def excel_to_dict(file_name, sheet):
    """
    Lee un excel y devuelve un diccionario
    """
    workbook = load_workbook(file_name)
    sheet = workbook[sheet]
    first_row = []  # The row where we stock the name of the column
    for col in range(1, sheet.max_column+1):
        if sheet.cell(row=1, column=col).value is None:
            continue
        first_row.append(sheet.cell(row=1, column=col).value)
    data = []
    for row in range(2, sheet.max_row+1):
        aux = 1

        elm = {}
        if sheet.cell(row=row, column=1).value is None:
            continue
        for col in range(1, sheet.max_column+1):
            if sheet.cell(row=1, column=col).value is None:
                continue
            elm[first_row[aux-1]] = sheet.cell(row=row, column=col).value
            aux += 1
        data.append(elm)
    print(data)
    return data


def comprobar_excel(data_excel, tabla='@ADV_EXPDATEMB'):
    """
    Comprueba que el excel esté bien formado
    """
    if ('DROP' in tabla.upper() or 'CREATE' in tabla.upper() or
       'ALTER' in tabla.upper() or 'TRUNCATE' in tabla.upper() or
        'DELETE' in tabla.upper() or 'INSERT' in tabla.upper() or
            'UPDATE' in tabla.upper() or 'SELECT' in tabla.upper()):
        raise (CustomError("PELIGRO DE INYECCIÓN SQL EN LA TABLA"))
    keyArray = []
    for key in data_excel[0]:
        keyArray.append(key)
