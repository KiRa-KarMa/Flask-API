from datetime import datetime
from openpyxl import load_workbook
from common.coms.hanaconnect import conectar_HanaDB
from ExcelToSAP.coms.hanaconnect import dictToSQLUpdate
import configparser
from common.encrypt.codeDecode import cargar_clave, desencriptar_items
import logging
import os
from time import sleep


class CustomError(Exception):
    """
    Excepción personalizada
    """
    pass


def borraArchivoTemp(filename):
    """
    Borra el archivo temporal

    :param filename: path del archivo temporal.
    """
    if os.path.isfile(filename):
        os.remove(filename)


def excel_to_hanadb(database):
    """
    Cuerpo del programa. Abrimos el archivo de excel, leemos los datos y
    actualizamos SAP.

    :return: Resultado de la ejecución en formato JSON.
    """
    parser = configparser.ConfigParser()
    parser.read('./config_code.ini')
    db = {}
    if parser.has_section('bd'):
        params = parser.items('bd')
        for param in params:
            if param[0] == 'db':
                db['db'] = param[1]
            else:
                db[param[0]] = (desencriptar_items(param[1],
                                                   cargar_clave()).decode(
                                                       "utf-8"))
    workbook = load_workbook('tempExcelToSAP.xlsx')
    sheets = workbook.sheetnames
    logging.info("Abrimos excel")

    for i in sheets:
        data_excel = excel_to_dict('tempExcelToSAP.xlsx', i)
        logging.info(f"Se cargan {len(data_excel)} registros del excel")
        with conectar_HanaDB(db['host'], db['user'], db['pass'],
                             db['port']) as conn:
            logging.info('Conectando a la BD')
            # Comprobaciones, si no esta OK return error
            comprobacion = comprobar_excel(data_excel, conn, database, i)
            if comprobacion[0] == -1:
                return comprobacion
            actualizados = []
            codigo = 0
            for row in data_excel:
                if 'Code' in row:
                    if row['Code'] is None:
                        logging.error(
                            'No existe el campo Code'
                        )
                        raise (CustomError("El campo Code no puede estar vacío"))
                    codigo = 1
                    dictToSQLUpdate(conn, row, i, database, 'Code')
                    logging.info(f'Actualizado {row["Code"]}')
                    print(f"El codigo {row['Code']} se ha actualizado.")
                    actualizados.append(row)

                elif 'U_AdvEnlECC' in row:
                    if row['U_AdvEnlECC'] is None:
                        logging.error(
                            'No existe el campo U_AdvEnlECC'
                        )
                        raise (CustomError("El campo U_AdvEnlECC no puede estar vacío"))
                    codigo = 1
                    logging.info(f"Actualizando {row['U_AdvEnlECC']}")
                    dictToSQLUpdate(conn, row, i, database, 'U_AdvEnlECC')
                    print(f"{row['U_AdvEnlECC']} se ha actualizado...")
                    actualizados.append(row)
            if codigo == 1:
                return [1, actualizados]


def excel_to_dict(file_name, sheet):
    """
    Lee un excel y devuelve un diccionario

    :param file_name: Path del Excel
    :param sheet: Hoja del Excel a convertir en diccionario
    :return: Diccionario con los datos del Excel
    """
    workbook = load_workbook(file_name)
    sheet = workbook[sheet]
    first_row = []  # The row where we stock the name of the column
    for col in range(1, sheet.max_column+1):
        if sheet.cell(row=2, column=col).value is None:
            continue
        first_row.append(sheet.cell(row=2, column=col).value)
    data = []
    for row in range(3, sheet.max_row+1):
        aux = 1

        elm = {}
        if sheet.cell(row=row, column=1).value is None:
            continue
        for col in range(1, sheet.max_column+1):
            if sheet.cell(row=1, column=col).value is None:
                continue
            elm[first_row[aux-1]] = sheet.cell(row=row, column=col).value
            aux += 1
        data.append(elm)
    print(data)
    return data


def comprobar_excel(data_excel, conn, database, tabla='@ADV_EXPDATEMB'):
    """
    Comprueba que el excel esté bien formado.

    :param data_excel: Diccionario con los datos del Excel
    :param conn: Conexión con HanaDB
    :param tabla: Tabla de la BD que necesitamos consultar, por defecto: @ADV_EXPDATEMB
    :return: Lista con los elementos, 1º codigo de devolucion que puede ser: {-1: Error, 1: OK}, 2º mensaje de respuesta
    """
    if ('DROP' in tabla.upper() or 'CREATE' in tabla.upper() or
        'ALTER' in tabla.upper() or 'TRUNCATE' in tabla.upper() or
        'DELETE' in tabla.upper() or 'INSERT' in tabla.upper() or
        'UPDATE' in tabla.upper() or 'SELECT' in tabla.upper()):
        msg = [-1, "PELIGRO DE INYECCIÓN SQL EN LA TABLA"]
        return msg
    if not (tabla.startswith('@')):
        msg = [-1, "No es una tabla definida por el usuario"]
        return msg
    keyArray = []
    for key in data_excel[0]:
        keyArray.append(key)
    tipos = obtenerTipos(conn, tabla, keyArray, database)
    if tipos[0] == -1:
        return tipos
    for i in data_excel:
        for j in tipos[1]:
            if (compruebaCelda(j['DATA_TYPE_NAME'], i[j['COLUMN_NAME']])):
                continue
            else:
                logging.info(f"El campo {j['COLUMN_NAME']} no es válido")
                msg = [-1, f"Error en la celda {j['COLUMN_NAME']}"]
    msg = [1]
    return msg


def obtenerTipos(conn, tabla, key_arr, database):
    """
    Obtiene los tipos de las columnas de una tabla

    :param conn: Conexión con HanaDB
    :param tabla: Tabla de la BD que necesitamos consultar
    :param key_arr: Lista con los valores de la columna Code de Excel 
    :return: Lista con los elementos, 1º codigo de devolucion que puede ser: {-1: Error, 1: OK}, 2º mensaje de respuesta
    """
    query = "SELECT COLUMN_NAME, DATA_TYPE_NAME"\
            " FROM TABLE_COLUMNS WHERE TABLE_NAME ="\
            f" '{tabla}' AND SCHEMA_NAME = '{database}' ORDER BY POSITION"
    curs = conn.cursor()
    curs.execute(query)
    df = curs.fetchall()
    res = []
    for i in df:
        if i['COLUMN_NAME'] in key_arr:
            res.append(i)
    for i in key_arr:
        if i not in [j['COLUMN_NAME'] for j in res]:
            msg = [-1, f"El campo {i} no está en la tabla"]
            return msg
    return [1, res]


def compruebaCelda(tipoEnBD, valorExcel):
    """
    Comprueba que el tipo de celda sea correcto.
    
    :param tipoEnBD: tipo de dato en la BD
    :param valorExcel: tipo de dato en Excel
    :return: True o False
    """
    logging.info(f"Comprobamos el valor {valorExcel}")

    if valorExcel is None:
        return True
    if (tipoEnBD in ['NCLOB', 'VARCHAR', 'NVARCHAR']):
        if (isinstance(valorExcel, str) or isinstance(valorExcel, int)):
            valorExcel = str(valorExcel)
            if ('DROP' in valorExcel.upper() or
                'CREATE' in valorExcel.upper() or
                'ALTER' in valorExcel.upper()
                or 'TRUNCATE' in valorExcel.upper() or
                'DELETE' in valorExcel.upper() or
                'INSERT' in valorExcel.upper()
                or 'UPDATE' in valorExcel.upper() or
                'SELECT' in valorExcel.upper()):
                logging.info(f"Peligro de inyección SQL en {valorExcel}")
                raise(CustomError("PELIGRO DE INYECCIÓN SQL EN LA TABLA"))
            return True
        else:
            return False
    elif (tipoEnBD == 'TIMESTAMP'):
        if (isinstance(valorExcel, datetime)):
            return True
        else:
            return False
    elif (tipoEnBD == 'INTEGER' or tipoEnBD == 'SMALLINT'):
        if (isinstance(valorExcel, int)):
            return True
        else:
            return False
    elif (tipoEnBD == 'DECIMAL'):
        if (isinstance(valorExcel, float) or isinstance(valorExcel, int)):
            return True
        else:
            return False
    else:
        return False
